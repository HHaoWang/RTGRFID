import numpy as np
import torch
from sklearn import metrics
from torch import nn
from openpyxl import Workbook

import Dataset
import utils
from RTGRFID import RTGRFID

dataDirectory = r"C:\Users\HHao\OneDrive\学习\研究生\小论文\实验数据\33"
userDataDirectory = r"C:\Users\HHao\OneDrive\学习\研究生\小论文\实验数据\User2"
numberOfRows = 3
numberOfCols = 3
tagArray2 = np.array([
    ['E280689400004013F1505D17', 'E280689400005013F1507117', 'E280689400005013F1506917'],
    ['E280689400004013F1507517', 'E280689400004013F155114C', 'E280689400005013F155154C'],
    ['E280689400004013F1506117', 'E280689400004013F1506D17', 'E280689400005013F1506517'],
])

tagArray1 = np.array([
    ['E280689400005013F1509918', 'E280689400005013F150B118', 'E280689400005013F150A518'],
    ['E280689400004013F1509D18', 'E280689400004013F150B518', 'E280689400005013F150A918'],
    ['E280689400005013F1509518', 'E280689400004013F150AD18', 'E280689400004013F150A118'],
])

seq_len = 90
epochs = 150

(train_data, train_labels), (valid_data, valid_labels), classes = Dataset.get_data(dataDirectory, tagArray1, tagArray2, seq_len,
                                                                                   0.3)
((user_train_data, user_train_labels), (user_valid_data, user_valid_labels),
 users_classes) = Dataset.get_data(userDataDirectory, tagArray1, tagArray2,
                                   seq_len, 0.3)
print("数据集准备完毕\n")

device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
results = []
acc = 0
model = RTGRFID(seq_len, len(classes), 2).to(device)
gesture_criterion = nn.CrossEntropyLoss().to(device)
user_criterion = nn.CrossEntropyLoss().to(device)
optimizer = torch.optim.Adam(model.parameters())

train_data_len = len(train_data)
if len(user_train_data) < train_data_len:
    train_data_len = len(user_train_data)

for epoch in range(epochs):
    idx = [i for i in range(train_data_len)]
    np.random.shuffle(idx)
    model.train()
    for i in range(len(idx)):
        p = float(i + epoch * train_data_len) / epochs / train_data_len
        alpha = 2. / (1. + np.exp(-10 * p)) - 1
        x_src = train_data[idx[i]]
        gestures_labels_true_src = torch.tensor(train_labels[idx[i]])
        user_labels_true_src = torch.tensor([1.0, 0.0])
        gestures_labels_pred_src, user_label_pred_src = model(x_src, alpha)
        gestures_loss_src = gesture_criterion(gestures_labels_pred_src, gestures_labels_true_src)
        user_loss_src = user_criterion(user_label_pred_src, user_labels_true_src)

        x_u = user_train_data[idx[i]]
        user_label_true_u = torch.tensor([0.0, 1.0]).to(device)
        _, user_label_pred_u = model(x_u, alpha)
        user_loss_u = user_criterion(user_label_pred_u, user_label_true_u)

        loss_all = gestures_loss_src + user_loss_src + user_loss_u
        optimizer.zero_grad()
        loss_all.backward()
        optimizer.step()

    model.eval()
    with torch.no_grad():
        y_true_labels_g = [utils.one_hot_to_string(label) for label in valid_labels]
        y_pred_proba_g = []
        y_pred_labels_g = []
        for i in range(len(valid_data)):
            x_src = valid_data[i]
            y = torch.tensor(valid_labels[i]).to(device)
            y_pred_g, y_pred_u = model(x_src, 1.0)
            y_pred_proba_g.append(y_pred_g)
            y_pred_labels_g.append(utils.one_hot_to_string(utils.convert_to_one_hot(y_pred_g)))

        accuracy_g = metrics.accuracy_score(y_true_labels_g, y_pred_labels_g)
        if accuracy_g > acc:
            torch.save(model, './model/bestModel.pth')
        results.append([])
        for (index, pred_label) in enumerate(y_pred_labels_g):
            if pred_label != y_true_labels_g[index]:
                results[epoch].append((y_true_labels_g[index], pred_label))
        print(f'Epoch:{epoch + 1:00},  Accuracy:{accuracy_g:.4f}')

        user_y_true_labels = [utils.one_hot_to_string(label) for label in user_valid_labels]
        user_y_pred_labels = []
        for i in range(len(user_valid_data)):
            x_src = user_valid_data[i]
            y = torch.tensor(user_valid_labels[i]).to(device)
            y_pred_g, y_pred_u = model(x_src, 1.0)
            user_y_pred_labels.append(utils.one_hot_to_string(utils.convert_to_one_hot(y_pred_g)))

        accuracy_g = metrics.accuracy_score(user_y_true_labels, user_y_pred_labels)
        print(f'\t\t User Predict Accuracy:{accuracy_g:.4f}')

print('训练结束，导出结果中...')

wb = Workbook()
ws = wb.active
ws.title = "说明"
ws['A1'] = "序号"
ws['B1'] = "名称"
for i, class_ in enumerate(classes):
    cell = ws.cell(row=i + 2, column=1)
    cell.value = i
    cell = ws.cell(row=i + 2, column=2)
    cell.value = class_
for index, result in enumerate(results):
    ws = wb.create_sheet(str(index + 1))
    ws['A1'] = "实际值"
    ws['B1'] = "预测值"
    for i, epoch_result in enumerate(result):
        cell = ws.cell(row=i + 2, column=1)
        cell.value = str(epoch_result[0])
        cell = ws.cell(row=i + 2, column=2)
        cell.value = str(epoch_result[1])
wb.save('结果.xlsx')
